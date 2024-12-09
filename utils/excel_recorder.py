import openpyxl
from datetime import datetime
import os
from utils.total_inspections import TotalInspectionsTracker

class ExcelRecorder:
    def __init__(self, filename="Latest_Detections.xlsx"):
        self.filename = filename
        self.total_tracker = TotalInspectionsTracker()
        self._ensure_file_exists()
        
    def _ensure_file_exists(self):
        if not os.path.exists(self.filename):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(["#", "Part Number", "Date", "Time", "Result"])
            wb.save(self.filename)
            
    def record_detection(self, part_number, part_status, clip_statuses):
        wb = openpyxl.load_workbook(self.filename)
        ws = wb.active
        
        # Get the last number used
        last_number = 1
        if ws.max_row > 1:
            last_number = ws.cell(row=2, column=1).value
            if last_number is None:
                last_number = 1
            else:
                try:
                    last_number = int(last_number)
                except ValueError:
                    last_number = 1
        
        # Prepare new row data
        now = datetime.now()
        result = "OK" if part_status and all(clip_statuses) else "NOK"
        
        # Insert new row after header
        ws.insert_rows(2)
        ws.cell(row=2, column=1, value=last_number + 1)
        ws.cell(row=2, column=2, value=part_number)
        ws.cell(row=2, column=3, value=now.strftime("%Y-%m-%d"))
        ws.cell(row=2, column=4, value=now.strftime("%H:%M:%S"))
        ws.cell(row=2, column=5, value=result)
        
        wb.save(self.filename)
        
        # Update total inspections
        self.total_tracker.update_inspections(part_number, result)