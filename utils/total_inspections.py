import openpyxl
import os
from typing import Dict, Tuple

class TotalInspectionsTracker:
    def __init__(self, filename="Total_Inspections.xlsx"):
        self.filename = filename
        self._ensure_file_exists()
        
    def _ensure_file_exists(self):
        if not os.path.exists(self.filename):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(["PN", "Total Inspections", "OK", "NOK"])
            wb.save(self.filename)
    
    def update_inspections(self, part_number: str, result: str) -> None:
        if part_number == "Unknown":
            return
            
        wb = openpyxl.load_workbook(self.filename)
        ws = wb.active
        
        # Find row for part number
        part_row = None
        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=1).value == part_number:
                part_row = row
                break
        
        # If part not found, add new row
        if part_row is None:
            part_row = ws.max_row + 1
            ws.cell(row=part_row, column=1, value=part_number)
            ws.cell(row=part_row, column=2, value=0)
            ws.cell(row=part_row, column=3, value=0)
            ws.cell(row=part_row, column=4, value=0)
        
        # Update counts
        total = ws.cell(row=part_row, column=2).value or 0
        ok_count = ws.cell(row=part_row, column=3).value or 0
        nok_count = ws.cell(row=part_row, column=4).value or 0
        
        ws.cell(row=part_row, column=2, value=total + 1)
        if result == "OK":
            ws.cell(row=part_row, column=3, value=ok_count + 1)
        else:
            ws.cell(row=part_row, column=4, value=nok_count + 1)
        
        wb.save(self.filename)
    
    def get_part_stats(self, part_number: str) -> Dict[str, str]:
        if not os.path.exists(self.filename) or part_number == "Unknown":
            return {
                "total": "0",
                "pass_rate": "0%"
            }
        
        wb = openpyxl.load_workbook(self.filename)
        ws = wb.active
        
        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=1).value == part_number:
                total = ws.cell(row=row, column=2).value or 0
                ok_count = ws.cell(row=row, column=3).value or 0
                
                pass_rate = (ok_count / total * 100) if total > 0 else 0
                
                return {
                    "total": str(total),
                    "pass_rate": f"{pass_rate:.1f}%"
                }
        
        return {
            "total": "0",
            "pass_rate": "0%"
        }